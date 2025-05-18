import os
from datetime import datetime
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from src.domain.entities.news import News
from src.domain.repositories.news_repository import NewsRepository
from src.infrastructure.logging.logger import logger

class ExcelNewsRepository(NewsRepository):
    def __init__(self, excel_path: str, images_dir: str):
        self.excel_path = excel_path
        self.images_dir = images_dir
        if not os.path.exists(images_dir):
            os.makedirs(images_dir)

    def _download_image(self, image_url: str, image_filename: str) -> str:
        """Download sincrono de imagem. Retorna o status do download."""
        if not image_url:
            return "URL não fornecida"

        try:
            logger.debug(f"Tentando baixar imagem de: {image_url}")
            response = requests.get(image_url)
            logger.debug(f"Status da resposta: {response.status_code}")
            
            if response.status_code == 200:
                image_path = os.path.join(self.images_dir, image_filename)
                with open(image_path, 'wb') as f:
                    f.write(response.content)
                logger.info(f"Imagem salva com sucesso em: {image_path}")
                return image_filename
            else:
                error_msg = f"Erro ao baixar imagem: {response.status_code}"
                logger.error(error_msg)
                return error_msg
        except Exception as e:
            error_msg = f"Erro ao salvar imagem: {str(e)}"
            logger.error(error_msg)
            return error_msg

    def save_news(self, news_list: list[News]) -> None:
        # Criar nova planilha
        wb = Workbook()
        ws = wb.active
        ws.title = "Notícias"

        # Definir estilos
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_alignment = Alignment(vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Definir cabeçalhos
        headers = ['Titulo', 'Data', 'Descricao', 'Imagem', 'Contagem da Frase', 'Contem Valor']

        # Aplicar estilos aos cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Download das imagens e preparação dos dados
        processed_news = []
        for news in news_list:
            if news.image_url and news.image_filename:
                image_status = self._download_image(news.image_url, news.image_filename)
                news.image_filename = image_status
            else:
                news.image_filename = "Sem imagem"
            processed_news.append(news)

        # Adicionar dados
        for row, news in enumerate(processed_news, 2):
            # Titulo
            cell = ws.cell(row=row, column=1)
            cell.value = news.title
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Data
            cell = ws.cell(row=row, column=2)
            cell.value = news.date.strftime('%Y-%m-%d %H:%M:%S')
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Descricao
            cell = ws.cell(row=row, column=3)
            cell.value = news.description
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Imagem
            cell = ws.cell(row=row, column=4)
            cell.value = news.image_filename
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Contagem da Frase
            cell = ws.cell(row=row, column=5)
            cell.value = news.search_phrase_count
            cell.alignment = cell_alignment
            cell.border = thin_border

            # Contem Valor
            cell = ws.cell(row=row, column=6)
            cell.value = "Sim" if news.has_money else "Nao"
            cell.alignment = cell_alignment
            cell.border = thin_border

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 100)  # Limita a largura máxima
            ws.column_dimensions[column_letter].width = adjusted_width

        # Congelar cabeçalho
        ws.freeze_panes = 'A2'

        # Salvar arquivo
        wb.save(self.excel_path)
        logger.info(f"Arquivo Excel salvo com sucesso em '{self.excel_path}'")

    def save_image(self, image_url: str, image_filename: str) -> None:
        """Metodo para compatibilidade com a interface."""
        self._download_image(image_url, image_filename) 